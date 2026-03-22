from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Callable

from django.core.cache import cache
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import transaction
from openpyxl import load_workbook
from rest_framework import status

from accesos.domain.services.email_domain_service import EmailDomainService
from .error_codes import ErrorCode
from .models import AprendizImportAudit, Sede, Usuario, sync_primary_membership

EXPECTED_COLUMNS_BASE = [
    "Nombres",
    "Apellidos",
    "Documento",
    "Telefono",
    "Correo",
    "Jornada",
    "Programa",
]
EXPECTED_COLUMNS_SUPERADMIN = [*EXPECTED_COLUMNS_BASE, "Sede"]

HEADER_ALIASES = {
    "nombres": "Nombres",
    "apellidos": "Apellidos",
    "documento": "Documento",
    "telefono": "Telefono",
    "correo": "Correo",
    "email": "Correo",
    "jornada": "Jornada",
    "programa": "Programa",
    "programaformacion": "Programa",
    "sede": "Sede",
}

DOCUMENT_RE = re.compile(r"^\d{6,30}$")
PHONE_RE = re.compile(r"^\d{7,15}$")
USERNAME_ALLOWED_RE = re.compile(r"[^a-z0-9._@+-]+")


@dataclass
class ImportValidationResult:
    rows: list[dict]
    errors: list[dict]
    duplicates_in_file: list[dict] = field(default_factory=list)


@dataclass
class ImportExecutionResult:
    created_count: int
    updated_count: int
    errors_count: int
    total_rows: int
    skipped_count: int = 0
    failed_count: int = 0
    row_results: list[dict] = field(default_factory=list)


class ImportServiceError(Exception):
    def __init__(
        self,
        *,
        code: str,
        message: str,
        status_code: int = status.HTTP_400_BAD_REQUEST,
        detail: dict | None = None,
        field: str | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code
        self.detail = detail
        self.field = field


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _ascii_token(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return text.strip().lower()


def _header_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _ascii_token(value))


def _normalize_program_token(value: str) -> str:
    token = re.sub(r"\s+", " ", _ascii_token(value))
    return token.strip()


def _normalize_phone(phone: str) -> str:
    digits = "".join(ch for ch in str(phone or "") if ch.isdigit())
    return digits


def _normalize_sede_token(value: str) -> str:
    return str(value or "").strip().lower().replace("_", "-").replace(" ", "-")


def _normalize_username_piece(value: str) -> str:
    raw = _ascii_token(value)
    first_token = next((piece for piece in re.split(r"\s+", raw) if piece), "")
    cleaned = USERNAME_ALLOWED_RE.sub("", first_token)
    return cleaned[:60]


def _build_username(first_name: str, last_name: str, documento: str) -> str:
    first = _normalize_username_piece(first_name)
    last = _normalize_username_piece(last_name)
    candidate = f"{first}.{last}".strip(".-_")
    if not candidate:
        candidate = f"aprendiz{documento[-6:]}"
    return candidate[:150]


def _username_with_document_suffix(base_username: str, documento: str) -> str:
    doc4 = "".join(ch for ch in str(documento or "") if ch.isdigit())[-4:]
    suffix = doc4 or "0000"
    candidate = f"{base_username}.{suffix}".strip(".-_")
    return candidate[:150]


def _resolve_import_username(
    base_username: str,
    documento: str,
    *,
    username_exists: Callable[[str], bool],
) -> str | None:
    candidate = base_username.strip()[:150]
    if candidate and not username_exists(candidate):
        return candidate

    candidate = _username_with_document_suffix(base_username, documento)
    if candidate and not username_exists(candidate):
        return candidate

    # Edge case: if base+doc4 also exists, keep deterministic fallback with a short numeric suffix.
    for idx in range(2, 100):
        attempt = f"{candidate}.{idx}"[:150]
        if not username_exists(attempt):
            return attempt
    return None


def _is_valid_email(email: str) -> bool:
    clean = email.lower().strip()
    try:
        validate_email(clean)
    except DjangoValidationError:
        return False
    return True


def _morning_code() -> str:
    for code, _label in Usuario.Jornada.choices:
        if _ascii_token(code).startswith("ma"):
            return code
    return "MANANA"


def _normalize_jornada(value: str) -> str:
    raw = _ascii_token(value).upper()
    morning = _morning_code()
    aliases = {
        _ascii_token(morning).upper(): morning,
        "MANANA": morning,
        "MAANA": morning,
        "MANA": morning,
        "TARDE": "TARDE",
        "NOCHE": "NOCHE",
    }
    return aliases.get(raw, raw)


def _build_program_lookup() -> dict[str, str]:
    out: dict[str, str] = {}
    programs = (
        Usuario.objects.exclude(programa_formacion__isnull=True)
        .exclude(programa_formacion__exact="")
        .values_list("programa_formacion", flat=True)
        .distinct()
    )
    for value in programs:
        normalized = _normalize_program_token(str(value))
        if normalized:
            out[normalized] = str(value).strip()
    return out


def _build_sede_lookup() -> dict[str, str]:
    out: dict[str, str] = {}
    for sede in Sede.objects.filter(is_active=True).only("code", "name"):
        out[_normalize_sede_token(sede.code)] = sede.code
        out[_normalize_sede_token(sede.name)] = sede.code
    return out


def _read_table_rows(content) -> list[list[str]]:
    name = str(getattr(content, "name", "") or "").lower()
    if hasattr(content, "seek"):
        content.seek(0)

    if name.endswith(".csv"):
        raw = content.read()
        if isinstance(raw, str):
            text = raw
        else:
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("latin-1")
        reader = csv.reader(io.StringIO(text))
        return [[_normalize_cell(col) for col in row] for row in reader]

    wb = load_workbook(filename=content, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_normalize_cell(value) for value in row])
    return rows


def _resolve_header_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw_header in enumerate(headers):
        token = _header_token(raw_header)
        canonical = HEADER_ALIASES.get(token)
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


def _sync_primary_membership_for_user(*, user: Usuario, role_code: str, sede: Sede | None):
    sync_primary_membership(
        user=user,
        role_code=role_code,
        sede=sede,
        is_active=True,
        can_switch_sede=role_code == Usuario.Rol.SUPERADMIN,
    )


def validate_excel(
    content,
    *,
    require_sede: bool = True,
    default_sede_code: str | None = None,
) -> ImportValidationResult:
    try:
        table_rows = _read_table_rows(content)
    except Exception:
        return ImportValidationResult(
            rows=[],
            errors=[
                {
                    "row": 1,
                    "code": "INVALID_FILE",
                    "message": "No se pudo leer el archivo. Verifica que sea Excel o CSV valido.",
                }
            ],
        )
    if not table_rows:
        return ImportValidationResult(
            rows=[],
            errors=[
                {
                    "row": 1,
                    "code": "EMPTY_FILE",
                    "message": "El archivo esta vacio.",
                }
            ],
        )

    headers = table_rows[0]
    header_map = _resolve_header_map(headers)
    required_headers = set(EXPECTED_COLUMNS_BASE + (["Sede"] if require_sede else []))
    missing_headers = [name for name in required_headers if name not in header_map]
    if missing_headers:
        return ImportValidationResult(
            rows=[],
            errors=[
                {
                    "row": 1,
                    "code": "INVALID_COLUMNS",
                    "message": "Faltan columnas requeridas para importar aprendices.",
                    "missing": sorted(missing_headers),
                    "expected": EXPECTED_COLUMNS_SUPERADMIN if require_sede else EXPECTED_COLUMNS_BASE,
                    "received": headers,
                }
            ],
        )

    sedes_lookup = _build_sede_lookup()
    sedes_by_code = {s.code: s for s in Sede.objects.filter(is_active=True).only("id", "code", "name")}
    if require_sede and not sedes_lookup:
        return ImportValidationResult(
            rows=[],
            errors=[
                {
                    "row": 1,
                    "code": "NO_ACTIVE_SEDES",
                    "message": "No hay sedes activas configuradas en el sistema.",
                    "field": "Sede",
                }
            ],
        )
    if not require_sede:
        if not default_sede_code:
            return ImportValidationResult(
                rows=[],
                errors=[
                    {
                        "row": 1,
                        "code": "MISSING_ADMIN_SEDE",
                        "message": "No se pudo determinar la sede del administrador para esta importacion.",
                        "field": "Sede",
                    }
                ],
            )
        if default_sede_code not in sedes_by_code:
            return ImportValidationResult(
                rows=[],
                errors=[
                    {
                        "row": 1,
                        "code": "INVALID_ADMIN_SEDE",
                        "message": "La sede del administrador no esta activa o no existe.",
                        "field": "Sede",
                    }
                ],
            )

    program_lookup = _build_program_lookup()
    has_program_catalog = bool(program_lookup)
    valid_jornadas = {choice for choice, _label in Usuario.Jornada.choices}

    rows: list[dict] = []
    errors: list[dict] = []
    doc_occurrences: dict[str, list[int]] = {}
    row_doc_map: dict[int, str] = {}

    def _get_value(values: list[str], key: str) -> str:
        idx = header_map[key]
        if idx >= len(values):
            return ""
        return _normalize_cell(values[idx])

    for row_num, raw_values in enumerate(table_rows[1:], start=2):
        nombres = _get_value(raw_values, "Nombres")
        apellidos = _get_value(raw_values, "Apellidos")
        documento = "".join(ch for ch in _get_value(raw_values, "Documento") if ch.isdigit())
        telefono = _normalize_phone(_get_value(raw_values, "Telefono"))
        correo = _get_value(raw_values, "Correo").strip().lower()
        jornada = _normalize_jornada(_get_value(raw_values, "Jornada"))
        programa_raw = _get_value(raw_values, "Programa")
        programa_norm = _normalize_program_token(programa_raw)

        row_errors: list[dict] = []
        if not nombres:
            row_errors.append(
                {
                    "row": row_num,
                    "code": "MISSING_REQUIRED",
                    "field": "Nombres",
                    "message": "Nombres es obligatorio.",
                }
            )
        if not apellidos:
            row_errors.append(
                {
                    "row": row_num,
                    "code": "MISSING_REQUIRED",
                    "field": "Apellidos",
                    "message": "Apellidos es obligatorio.",
                }
            )
        if not documento:
            row_errors.append(
                {
                    "row": row_num,
                    "code": "MISSING_REQUIRED",
                    "field": "Documento",
                    "message": "Documento es obligatorio.",
                }
            )
        elif not DOCUMENT_RE.fullmatch(documento):
            row_errors.append(
                {
                    "row": row_num,
                    "code": "INVALID_DOCUMENTO",
                    "field": "Documento",
                    "message": "Documento invalido. Debe tener al menos 6 digitos.",
                }
            )
        else:
            doc_occurrences.setdefault(documento, []).append(row_num)
            row_doc_map[row_num] = documento
        if telefono and not PHONE_RE.fullmatch(telefono):
            row_errors.append(
                {
                    "row": row_num,
                    "code": "INVALID_TELEFONO",
                    "field": "Telefono",
                    "message": "Telefono invalido.",
                }
            )
        if correo and not _is_valid_email(correo):
            row_errors.append(
                {
                    "row": row_num,
                    "code": "INVALID_EMAIL",
                    "field": "Correo",
                    "message": "Correo invalido.",
                }
            )
        if jornada not in valid_jornadas:
            row_errors.append(
                {
                    "row": row_num,
                    "code": "INVALID_JORNADA",
                    "field": "Jornada",
                    "message": f"Jornada invalida: {jornada or '(vacia)'}",
                }
            )
        if not programa_norm:
            row_errors.append(
                {
                    "row": row_num,
                    "code": "MISSING_REQUIRED",
                    "field": "Programa",
                    "message": "Programa es obligatorio.",
                }
            )
        elif has_program_catalog and programa_norm not in program_lookup:
            row_errors.append(
                {
                    "row": row_num,
                    "code": "INVALID_PROGRAMA",
                    "field": "Programa",
                    "message": f"Programa invalido: {programa_raw}",
                }
            )

        if require_sede:
            sede_source = _get_value(raw_values, "Sede")
            sede_code = sedes_lookup.get(_normalize_sede_token(sede_source))
            if not sede_code:
                row_errors.append(
                    {
                        "row": row_num,
                        "code": "INVALID_SEDE",
                        "field": "Sede",
                        "message": "Sede invalida. Debe coincidir con una sede activa.",
                    }
                )
        else:
            sede_code = str(default_sede_code or "").strip()

        if correo and sede_code:
            domain_check = EmailDomainService.validate(
                email=correo,
                role_code=Usuario.Rol.APRENDIZ,
                sede=sedes_by_code.get(sede_code),
            )
            if not domain_check.allowed:
                row_errors.append(
                    {
                        "row": row_num,
                        "code": domain_check.code or ErrorCode.EMAIL_DOMAIN_NOT_ALLOWED,
                        "field": "Correo",
                        "message": domain_check.message or "Dominio de correo no permitido para esta sede.",
                    }
                )

        if row_errors:
            errors.extend(row_errors)
            continue

        program_value = program_lookup.get(programa_norm, programa_raw.strip())
        username_sugerido = _build_username(nombres, apellidos, documento)
        rows.append(
            {
                "source_row": row_num,
                "first_name": nombres,
                "last_name": apellidos,
                "documento": documento,
                "telefono": telefono,
                "email": correo,
                "jornada": jornada,
                "programa_formacion": program_value,
                "sede_principal": sede_code,
                "username_sugerido": username_sugerido,
            }
        )

    duplicate_docs = {doc for doc, rows_found in doc_occurrences.items() if len(rows_found) > 1}
    if duplicate_docs:
        filtered_rows: list[dict] = []
        duplicate_errors: list[dict] = []
        duplicate_row_numbers = {row_num for row_num, document in row_doc_map.items() if document in duplicate_docs}
        existing_duplicate_rows = {
            int(err.get("row") or 0) for err in errors if str(err.get("code") or "").upper() == "DUPLICATE_IN_FILE"
        }
        emitted_duplicate_rows = set(existing_duplicate_rows)
        for row in rows:
            if int(row.get("source_row") or 0) not in duplicate_row_numbers:
                filtered_rows.append(row)
                continue
            duplicate_errors.append(
                {
                    "row": row["source_row"],
                    "code": "DUPLICATE_IN_FILE",
                    "field": "Documento",
                    "documento": row["documento"],
                    "message": "Documento duplicado dentro del archivo.",
                }
            )
            emitted_duplicate_rows.add(int(row.get("source_row") or 0))
        for row_num in sorted(duplicate_row_numbers):
            if row_num in emitted_duplicate_rows:
                continue
            documento = row_doc_map.get(row_num, "")
            if not documento:
                continue
            duplicate_errors.append(
                {
                    "row": row_num,
                    "code": "DUPLICATE_IN_FILE",
                    "field": "Documento",
                    "documento": documento,
                    "message": "Documento duplicado dentro del archivo.",
                }
            )
        rows = filtered_rows
        errors.extend(duplicate_errors)
        duplicates_in_file = duplicate_errors
    else:
        duplicates_in_file = []

    errors.sort(key=lambda item: (int(item.get("row", 0)), str(item.get("field", ""))))
    return ImportValidationResult(
        rows=rows,
        errors=errors,
        duplicates_in_file=duplicates_in_file,
    )


def cache_import_payload(import_id: str, user_id: int, rows: list[dict], errors: list[dict], ttl_sec: int = 15 * 60):
    cache.set(f"sadi:import:{import_id}", {"user_id": user_id, "rows": rows, "errors": errors}, timeout=ttl_sec)


def get_cached_import_payload(import_id: str):
    return cache.get(f"sadi:import:{import_id}")


def execute_aprendices_import(
    *,
    rows: list[dict],
    imported_by: Usuario,
    errors: list[dict] | None = None,
) -> ImportExecutionResult:
    created_count = 0
    updated_count = 0
    skipped_count = 0
    failed_count = 0
    pre_validation_errors = errors or []
    row_results: list[dict] = []
    reserved_usernames: set[str] = set()

    try:
        with transaction.atomic():
            sede_codes = [
                str(r.get("sede_principal", "")).strip() for r in rows if str(r.get("sede_principal", "")).strip()
            ]
            sedes_by_code = {s.code: s for s in Sede.objects.filter(code__in=sede_codes)}
            existing_users = {
                str(item["documento"]): item
                for item in Usuario.objects.filter(
                    documento__in=[str(r.get("documento", "")).strip() for r in rows]
                ).values(
                    "documento",
                    "first_name",
                    "last_name",
                    "sede_principal__name",
                )
            }
            existing_docs = set(existing_users.keys())

            def _username_exists(candidate: str) -> bool:
                key = str(candidate or "").strip().lower()
                if not key:
                    return True
                if key in reserved_usernames:
                    return True
                exists = Usuario.objects.filter(username__iexact=candidate).exists()
                if exists:
                    reserved_usernames.add(key)
                return exists

            for row in rows:
                source_row = int(row.get("source_row") or 0)
                documento = str(row.get("documento", "")).strip()
                sede_code = str(row.get("sede_principal", "")).strip()
                username = str(row.get("username_sugerido", "")).strip() or _build_username(
                    str(row.get("first_name", "")),
                    str(row.get("last_name", "")),
                    documento,
                )
                username = _resolve_import_username(username, documento, username_exists=_username_exists)
                if not username:
                    raise ImportServiceError(
                        code="USERNAME_EXISTS",
                        message="No se pudo asignar un username unico para el aprendiz.",
                        field="Nombres",
                        detail={"row": source_row, "documento": documento},
                    )

                if documento in existing_docs:
                    existing = existing_users.get(documento) or {}
                    existing_name = f"{existing.get('first_name', '')} {existing.get('last_name', '')}".strip()
                    skipped_count += 1
                    row_results.append(
                        {
                            "row": source_row,
                            "documento": documento,
                            "status": "skipped",
                            "code": ErrorCode.DOCUMENT_EXISTS,
                            "field": "Documento",
                            "reason": "Documento ya existe en el sistema.",
                            "username_asignado": None,
                            "existing_nombre": existing_name or None,
                            "existing_sede": existing.get("sede_principal__name"),
                        }
                    )
                    continue

                sede_obj = sedes_by_code.get(sede_code)
                if not sede_obj:
                    raise ImportServiceError(
                        code="INVALID_SEDE",
                        message="Sede invalida para la fila importada.",
                        field="Sede",
                        detail={"row": source_row, "documento": documento, "sede_principal": sede_code},
                    )

                email = str(row.get("email", "")).strip().lower()
                if email:
                    domain_check = EmailDomainService.validate(
                        email=email,
                        role_code=Usuario.Rol.APRENDIZ,
                        sede=sede_obj,
                    )
                    if not domain_check.allowed:
                        raise ImportServiceError(
                            code=domain_check.code or ErrorCode.EMAIL_DOMAIN_NOT_ALLOWED,
                            message=domain_check.message or "Dominio de correo no permitido para esta sede.",
                            field="Correo",
                            detail={"row": source_row, "documento": documento, "email": email},
                        )

                created = Usuario.objects.create(
                    username=username,
                    first_name=str(row.get("first_name", "")).strip(),
                    last_name=str(row.get("last_name", "")).strip(),
                    email=email,
                    telefono=str(row.get("telefono", "")).strip(),
                    documento=documento,
                    sede_principal=sede_obj,
                    jornada=str(row.get("jornada", "")).strip(),
                    programa_formacion=str(row.get("programa_formacion", "")).strip(),
                    rol=Usuario.Rol.APRENDIZ,
                    estado=Usuario.Estado.ACTIVO,
                    must_change_password=True,
                )
                initial_password = documento[-6:]
                created.set_password(initial_password)
                created.save(update_fields=["password"])
                _sync_primary_membership_for_user(
                    user=created,
                    role_code=Usuario.Rol.APRENDIZ,
                    sede=sede_obj,
                )
                existing_docs.add(documento)
                reserved_usernames.add(str(username).lower())
                created_count += 1
                row_results.append(
                    {
                        "row": source_row,
                        "documento": documento,
                        "status": "created",
                        "code": None,
                        "field": None,
                        "reason": "Aprendiz creado correctamente.",
                        "username_asignado": username,
                    }
                )

            AprendizImportAudit.objects.create(
                imported_by=imported_by,
                total_rows=len(rows) + len(pre_validation_errors),
                created_count=created_count,
                updated_count=updated_count,
                errors_count=len(pre_validation_errors) + skipped_count + failed_count,
            )

        return ImportExecutionResult(
            created_count=created_count,
            updated_count=updated_count,
            errors_count=len(pre_validation_errors) + skipped_count + failed_count,
            total_rows=len(rows) + len(pre_validation_errors),
            skipped_count=skipped_count,
            failed_count=failed_count,
            row_results=row_results,
        )
    except ImportServiceError:
        raise
    except Exception as exc:
        raise ImportServiceError(
            code=ErrorCode.SERVER_ERROR,
            message="No se pudo completar la importacion masiva. Intenta nuevamente.",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"reason": "unexpected_error"},
        ) from exc
